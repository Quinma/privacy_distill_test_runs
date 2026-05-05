#!/usr/bin/env python3
import json
import math
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook


REPO = Path(__file__).resolve().parents[1]
ROOT = REPO.parent
LOCAL = ROOT / "local_repo"
EXP = ROOT / "exp"
WORKBOOK = EXP / "outputs" / "score_summary.xlsx"
sys.path.insert(0, str(REPO / "src"))
from compute_stats import bootstrap_ci_diff, paired_permutation_test, sign_test

RUN_TAG = "gpt-neo-1.3b-local"
SCALE = "neo-1.3B"
MIA_DIR = LOCAL / "outputs" / RUN_TAG / "mia"
MATCHED_DIR = LOCAL / "outputs" / RUN_TAG / "mia_matched"
SEED_ROOT = LOCAL / "outputs" / RUN_TAG / "seed_reps"
PYTHIA_MIA_DIR = EXP / "outputs" / "pythia-1.4b" / "mia"
PYTHIA_ATTACK_DIR = EXP / "outputs" / "pythia-1.4b"
PYTHIA_PLACEBO_DIR = EXP / "outputs" / "pythia-1.4b-placebo-npo-s13"
PYTHIA_SEED_ROOT = EXP / "outputs" / "pythia-1.4b" / "seed_reps"
NEO_REPAIR_1P3B_DIR = EXP / "outputs" / "myriad_c23_repair_20260421" / "outputs" / "gpt-neo-1.3b-local"
NEO_PLACEBO_1P3B_DIR = EXP / "outputs" / "myriad_placebo_neo13_20260426" / "outputs" / "gpt-neo-1.3b-local-placebo-npo-s13"
SEEDED_PLACEBO_STAGE_DIR = EXP / "outputs" / "myriad_seed_placebo_c1_20260428" / "outputs"
PYTHIA_SEEDED_PLACEBO_DIR = SEEDED_PLACEBO_STAGE_DIR / "pythia-1.4b"
NEO_SEEDED_PLACEBO_DIR = SEEDED_PLACEBO_STAGE_DIR / "gpt-neo-1.3b-local"
MANUAL_INGEST_DIR = EXP / "outputs" / "manual_ingest"
NEO_2P7B_MANUAL_STATS = MANUAL_INGEST_DIR / "neo_2p7b_stats_20260417.json"
NEO_C6_MANUAL_STATS = MANUAL_INGEST_DIR / "neo_c6_20260418.json"
NEO_SEED_C6_MANUAL_STATS = MANUAL_INGEST_DIR / "neo_seed_c6_manual_20260418.json"
SEED_TEACHER_ATTACK_SHEET = "seed_teacher_del_att_1.4B"
SEED_TEACHER_SUMMARY_SHEET = "seed_teacher_del_sum_1.4B"
FOLLOWUP_DIR = EXP / "outputs" / "attack_followups"


def read_json(path: Path):
    with path.open() as f:
        return json.load(f)


def norm(value):
    if isinstance(value, float) and math.isinf(value):
        return "inf"
    return value


def ensure_sheet(wb, title, header):
    if title in wb.sheetnames:
        ws = wb[title]
        if ws.max_row == 0:
            ws.append(header)
        elif [c.value for c in ws[1]] != header:
            ws.delete_rows(1, ws.max_row)
            ws.append(header)
        return ws
    ws = wb.create_sheet(title)
    ws.append(header)
    return ws


def rename_sheet_if_present(wb, old_title, new_title):
    if old_title in wb.sheetnames and new_title not in wb.sheetnames:
        wb[old_title].title = new_title


def upsert(ws, key_cols, row):
    header = [c.value for c in ws[1]]
    key_idx = [header.index(k) for k in key_cols]
    row_values = [row.get(col) for col in header]
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, i + 1).value for i in key_idx]
        if vals == [row.get(k) for k in key_cols]:
            for c, val in enumerate(row_values, start=1):
                ws.cell(r, c).value = norm(val)
            return
    ws.append([norm(v) for v in row_values])


def sort_sheet(ws, sort_keys):
    header = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in r):
            rows.append(dict(zip(header, r)))
    rows.sort(key=lambda row: tuple("" if row.get(k) is None else str(row.get(k)) for k in sort_keys))
    ws.delete_rows(2, max(0, ws.max_row - 1))
    for row in rows:
        ws.append([row.get(col) for col in header])


def delete_matching_rows(ws, predicate):
    for r in range(ws.max_row, 1, -1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if predicate(row):
            ws.delete_rows(r, 1)


def direct_mia_rows():
    stats = read_json(MIA_DIR / "stats_bootstrap.json")
    cond_map = {
        "C1": ("C1 student", MIA_DIR / "c1_student.json"),
        "C2": ("C2 student", MIA_DIR / "c2_student.json"),
        "C3": ("C3 student", MIA_DIR / "c3_student.json"),
        "C4": ("C4 teacher", MIA_DIR / "c4_teacher.json"),
        "C5": ("C5 student", MIA_DIR / "c5_student.json"),
        "C5r": ("C5r student", MIA_DIR / "c5r_student.json"),
    }
    rows = []
    for key, (label, path) in cond_map.items():
        if key not in stats["bootstrap"] or not path.exists():
            continue
        entry = stats["bootstrap"][key]
        rows.append({
            "scale": SCALE,
            "condition": label,
            "auroc": entry["auroc"],
            "num_companies": entry["n_companies"],
            "status": "ok",
            "file": str(path),
        })

    for label, path in [
        ("C5 teacher", MIA_DIR / "c5_teacher.json"),
        ("C5r teacher", MIA_DIR / "c5r_teacher.json"),
        ("C6 teacher", MIA_DIR / "c6_teacher.json"),
        ("C6 student", MIA_DIR / "c6_student.json"),
    ]:
        if not path.exists():
            continue
        data = read_json(path)
        rows.append({
            "scale": SCALE,
            "condition": label,
            "auroc": data["auroc"],
            "num_companies": len(data["per_company"]),
            "status": "ok",
            "file": str(path),
        })
    return rows


def paired_rows():
    stats = read_json(MIA_DIR / "stats_bootstrap.json")["paired"]
    rows = []
    for comparison, data in stats.items():
        rows.append({
            "scale": SCALE,
            "comparison": comparison,
            "mean_diff": data["mean_diff"],
            "ci_low": data["bootstrap_ci_diff"][0],
            "ci_high": data["bootstrap_ci_diff"][1],
            "sign_p": data["sign_test"]["p_value_two_sided"],
            "sign_pos": data["sign_test"]["pos"],
            "sign_neg": data["sign_test"]["neg"],
            "perm_p": data["perm_test"]["p_value_two_sided"],
            "n": data["n"],
        })
    return rows


def utility_rows(kind):
    dataset = "distill" if kind == "distill" else "eval_target_holdout"
    suffix = "" if kind == "distill" else "_holdout"
    mapping = [
        ("c1_student", MIA_DIR / f"utility_c1_student{suffix}.json"),
        ("c2_student", MIA_DIR / f"utility_c2_student{suffix}.json"),
        ("c3_student", MIA_DIR / f"utility_c3_student{suffix}.json"),
        ("c4_teacher", MIA_DIR / f"utility_c4_teacher{suffix}.json"),
        ("c5_student", MIA_DIR / f"utility_c5_student{suffix}.json"),
        ("c5_teacher", MIA_DIR / f"utility_c5_teacher{suffix}.json"),
        ("c5r_student", MIA_DIR / f"utility_c5r_student{suffix}.json"),
        ("c5r_teacher", MIA_DIR / f"utility_c5r_teacher{suffix}.json"),
        ("c6_student", MIA_DIR / f"utility_c6_student{suffix}.json"),
        ("c6_teacher", MIA_DIR / f"utility_c6_teacher{suffix}.json"),
    ]
    rows = []
    for label, path in mapping:
        if not path.exists():
            continue
        data = read_json(path)
        rows.append({
            "scale": SCALE,
            "model": f"{label}{'_holdout' if kind == 'holdout' else ''}",
            "mean_loss": data["mean_loss"],
            "perplexity": norm(data["perplexity"]),
            "num_examples": data["num_examples"],
            "dataset": data["dataset"],
            "file": str(path),
        })
    return rows


def matched_rows():
    stats_path = MATCHED_DIR / "stats_bootstrap.json"
    if not stats_path.exists():
        return [], []
    stats = read_json(stats_path)
    mia_rows = []
    for cond, data in stats["bootstrap"].items():
        mia_rows.append({
            "condition": cond,
            "auroc": data["auroc"],
            "auroc_ci95_low": data["auroc_ci95"][0],
            "auroc_ci95_high": data["auroc_ci95"][1],
            "auroc_doc": data["auroc_doc"],
            "mean_loss_ratio": data["mean_loss_ratio"],
            "loss_ratio_ci95_low": data["loss_ratio_ci95"][0],
            "loss_ratio_ci95_high": data["loss_ratio_ci95"][1],
            "n_companies": data["n_companies"],
            "n_nonmember_companies": data["n_nonmember_companies"],
        })
    paired_rows_out = []
    for comparison, data in stats["paired"].items():
        paired_rows_out.append({
            "scale": SCALE,
            "comparison": comparison,
            "mean_diff": data["mean_diff"],
            "ci_low": data["bootstrap_ci_diff"][0],
            "ci_high": data["bootstrap_ci_diff"][1],
            "sign_p": data["sign_test"]["p_value_two_sided"],
            "sign_pos": data["sign_test"]["pos"],
            "sign_neg": data["sign_test"]["neg"],
            "perm_p": data["perm_test"]["p_value_two_sided"],
            "n": data["n"],
        })
    return mia_rows, paired_rows_out


def neo_2p7b_manual_rows():
    if not NEO_2P7B_MANUAL_STATS.exists():
        return [], [], [], [], [], []

    stats = read_json(NEO_2P7B_MANUAL_STATS)
    cond_map = {
        "C1": "C1 student",
        "C2": "C2 student",
        "C3": "C3 student",
        "C4": "C4 teacher",
    }

    mia_rows = []
    for cond, entry in stats["canonical"]["bootstrap"].items():
        if cond not in cond_map:
            continue
        mia_rows.append({
            "scale": "neo-2.7B",
            "condition": cond_map[cond],
            "auroc": entry["auroc"],
            "num_companies": entry["n_companies"],
            "status": "manual_myria_ingest",
            "file": str(NEO_2P7B_MANUAL_STATS),
        })

    paired_rows_out = []
    for comparison, data in stats["canonical"]["paired"].items():
        paired_rows_out.append({
            "scale": "neo-2.7B",
            "comparison": comparison,
            "mean_diff": data["mean_diff"],
            "ci_low": data["bootstrap_ci_diff"][0],
            "ci_high": data["bootstrap_ci_diff"][1],
            "sign_p": data["sign_test"]["p_value_two_sided"],
            "sign_pos": data["sign_test"]["pos"],
            "sign_neg": data["sign_test"]["neg"],
            "perm_p": data["perm_test"]["p_value_two_sided"],
            "n": data["n"],
        })

    matched_mia_rows = []
    for cond, entry in stats["matched"]["bootstrap"].items():
        matched_mia_rows.append({
            "condition": cond,
            "auroc": entry["auroc"],
            "auroc_ci95_low": entry["auroc_ci95"][0],
            "auroc_ci95_high": entry["auroc_ci95"][1],
            "auroc_doc": entry["auroc_doc"],
            "mean_loss_ratio": entry["mean_loss_ratio"],
            "loss_ratio_ci95_low": entry["loss_ratio_ci95"][0],
            "loss_ratio_ci95_high": entry["loss_ratio_ci95"][1],
            "n_companies": entry["n_companies"],
            "n_nonmember_companies": entry["n_nonmember_companies"],
        })

    paired_matched_rows_out = []
    for comparison, data in stats["matched"]["paired"].items():
        paired_matched_rows_out.append({
            "scale": "neo-2.7B",
            "comparison": comparison,
            "mean_diff": data["mean_diff"],
            "ci_low": data["bootstrap_ci_diff"][0],
            "ci_high": data["bootstrap_ci_diff"][1],
            "sign_p": data["sign_test"]["p_value_two_sided"],
            "sign_pos": data["sign_test"]["pos"],
            "sign_neg": data["sign_test"]["neg"],
            "perm_p": data["perm_test"]["p_value_two_sided"],
            "n": data["n"],
        })

    utility_distill_rows = []
    for model, data in stats.get("utility", {}).get("distill", {}).items():
        utility_distill_rows.append({
            "scale": "neo-2.7B",
            "model": model,
            "mean_loss": data["mean_loss"],
            "perplexity": norm(data["perplexity"]),
            "num_examples": data["num_examples"],
            "dataset": data["dataset"],
            "file": str(NEO_2P7B_MANUAL_STATS),
        })

    utility_holdout_rows = []
    for model, data in stats.get("utility", {}).get("holdout", {}).items():
        utility_holdout_rows.append({
            "scale": "neo-2.7B",
            "model": model,
            "mean_loss": data["mean_loss"],
            "perplexity": norm(data["perplexity"]),
            "num_examples": data["num_examples"],
            "dataset": data["dataset"],
            "file": str(NEO_2P7B_MANUAL_STATS),
        })

    return mia_rows, paired_rows_out, matched_mia_rows, paired_matched_rows_out, utility_distill_rows, utility_holdout_rows


def neo_c6_manual_rows():
    if not NEO_C6_MANUAL_STATS.exists():
        return [], [], [], []

    stats = read_json(NEO_C6_MANUAL_STATS)

    mia_rows = []
    for key, label in [("c6_teacher", "C6 teacher"), ("c6_student", "C6 student")]:
        entry = stats.get("mia", {}).get(key)
        if not entry:
            continue
        mia_rows.append({
            "scale": SCALE,
            "condition": label,
            "auroc": entry["auroc"],
            "num_companies": entry["num_companies"],
            "status": "manual_myriad_ingest",
            "file": str(NEO_C6_MANUAL_STATS),
        })

    utility_distill_rows = []
    for key in ["c6_teacher", "c6_student"]:
        entry = stats.get("utility", {}).get("distill", {}).get(key)
        if not entry:
            continue
        utility_distill_rows.append({
            "scale": SCALE,
            "model": key,
            "mean_loss": entry["mean_loss"],
            "perplexity": norm(entry["perplexity"]),
            "num_examples": entry["num_examples"],
            "dataset": entry["dataset"],
            "file": str(NEO_C6_MANUAL_STATS),
        })

    utility_holdout_rows = []
    for key in ["c6_teacher_holdout", "c6_student_holdout"]:
        entry = stats.get("utility", {}).get("holdout", {}).get(key)
        if not entry:
            continue
        utility_holdout_rows.append({
            "scale": SCALE,
            "model": key,
            "mean_loss": entry["mean_loss"],
            "perplexity": norm(entry["perplexity"]),
            "num_examples": entry["num_examples"],
            "dataset": entry["dataset"],
            "file": str(NEO_C6_MANUAL_STATS),
        })

    run_metrics_rows = []
    train_metrics = stats.get("train_metrics", {})
    if train_metrics:
        run_metrics_rows.append({
            "scale": SCALE,
            "condition": "C6 teacher NPO",
            "seed": train_metrics.get("seed"),
            "method": train_metrics.get("method"),
            "total_steps": train_metrics.get("total_steps"),
            "initial_forget_loss": train_metrics.get("initial_forget_loss"),
            "final_forget_loss": train_metrics.get("final_forget_loss"),
            "initial_retain_loss": train_metrics.get("initial_retain_loss"),
            "final_retain_loss": train_metrics.get("final_retain_loss"),
            "final_forget_log_ratio": train_metrics.get("final_forget_log_ratio"),
            "ref_checksum_match": train_metrics.get("ref_checksum_match"),
            "file": str(NEO_C6_MANUAL_STATS),
        })

    return mia_rows, utility_distill_rows, utility_holdout_rows, run_metrics_rows


def neo_seed_c6_manual_rows():
    if not NEO_SEED_C6_MANUAL_STATS.exists():
        return [], [], [], []

    stats = read_json(NEO_SEED_C6_MANUAL_STATS)
    mia_rows = []
    utility_distill_rows = []
    utility_holdout_rows = []
    run_metrics_rows = []

    for seed_str, seed_data in sorted(stats.get("seeds", {}).items(), key=lambda kv: int(kv[0])):
        seed = int(seed_str)
        for key in ["c6_teacher", "c6_student"]:
            entry = seed_data.get("mia", {}).get(key)
            if not entry:
                continue
            mia_rows.append({
                "seed": seed,
                "condition": key,
                "auroc": entry["auroc"],
                "auroc_doc": entry["auroc_doc"],
                "num_companies": entry["num_companies"],
                "file": str(NEO_SEED_C6_MANUAL_STATS),
            })

        for key in ["c6_teacher", "c6_student"]:
            entry = seed_data.get("utility", {}).get("distill", {}).get(key)
            if not entry:
                continue
            utility_distill_rows.append({
                "seed": seed,
                "model": key,
                "mean_loss": entry["mean_loss"],
                "perplexity": norm(entry["perplexity"]),
                "num_examples": entry["num_examples"],
                "dataset": entry["dataset"],
                "file": str(NEO_SEED_C6_MANUAL_STATS),
            })

        for key in ["c6_teacher_holdout", "c6_student_holdout"]:
            entry = seed_data.get("utility", {}).get("holdout", {}).get(key)
            if not entry:
                continue
            utility_holdout_rows.append({
                "seed": seed,
                "model": key,
                "mean_loss": entry["mean_loss"],
                "perplexity": norm(entry["perplexity"]),
                "num_examples": entry["num_examples"],
                "dataset": entry["dataset"],
                "file": str(NEO_SEED_C6_MANUAL_STATS),
            })

        train_metrics = seed_data.get("train_metrics", {})
        if train_metrics:
            run_metrics_rows.append({
                "seed": seed,
                "method": train_metrics.get("method"),
                "total_steps": train_metrics.get("total_steps"),
                "initial_forget_loss": train_metrics.get("initial_forget_loss"),
                "final_forget_loss": train_metrics.get("final_forget_loss"),
                "initial_retain_loss": train_metrics.get("initial_retain_loss"),
                "final_retain_loss": train_metrics.get("final_retain_loss"),
                "final_forget_log_ratio": train_metrics.get("final_forget_log_ratio"),
                "ref_checksum_match": train_metrics.get("ref_checksum_match"),
                "file": str(NEO_SEED_C6_MANUAL_STATS),
            })

    return mia_rows, utility_distill_rows, utility_holdout_rows, run_metrics_rows


def seed_rows():
    mia_rows = []
    util_distill_rows = []
    util_holdout_rows = []
    for seed_dir in sorted(SEED_ROOT.glob("seed_*")):
        try:
            seed = int(seed_dir.name.split("_", 1)[1])
        except Exception:
            continue
        mia_dir = seed_dir / "mia"
        for name in ["c1_student", "c3_student", "c5m_student", "c5m_teacher", "c5r_student", "c5r_teacher", "c6_student", "c6_teacher"]:
            path = mia_dir / f"{name}.json"
            if path.exists():
                data = read_json(path)
                mia_rows.append({
                    "seed": seed,
                    "condition": name,
                    "auroc": data["auroc"],
                    "auroc_doc": data["auroc_doc"],
                    "num_companies": len(data["per_company"]),
                    "file": str(path),
                })
        for name in [
            "utility_c1_student",
            "utility_c3_student",
            "utility_c5m_student",
            "utility_c5m_teacher",
            "utility_c5r_student",
            "utility_c5r_teacher",
            "utility_c6_student",
            "utility_c6_teacher",
        ]:
            path = mia_dir / f"{name}.json"
            if path.exists():
                data = read_json(path)
                util_distill_rows.append({
                    "seed": seed,
                    "model": name.replace("utility_", ""),
                    "mean_loss": data["mean_loss"],
                    "perplexity": norm(data["perplexity"]),
                    "num_examples": data["num_examples"],
                    "dataset": data["dataset"],
                    "file": str(path),
                })
        for name in [
            "utility_c1_student_holdout",
            "utility_c3_student_holdout",
            "utility_c5m_student_holdout",
            "utility_c5m_teacher_holdout",
            "utility_c5r_student_holdout",
            "utility_c5r_teacher_holdout",
            "utility_c6_student_holdout",
            "utility_c6_teacher_holdout",
        ]:
            path = mia_dir / f"{name}.json"
            if path.exists():
                data = read_json(path)
                util_holdout_rows.append({
                    "seed": seed,
                    "model": name.replace("utility_", ""),
                    "mean_loss": data["mean_loss"],
                    "perplexity": norm(data["perplexity"]),
                    "num_examples": data["num_examples"],
                    "dataset": data["dataset"],
                    "file": str(path),
                })
    return mia_rows, util_distill_rows, util_holdout_rows


def pythia_seed_c5_teacher_backfill_rows():
    root = Path("/path/to/privacy_distill_test_runs")
    dataset = str(root / "data" / "datasets" / "pythia-1.4b" / "distill")
    rows = [
        {
            "seed": 13,
            "model": "c5_teacher",
            "mean_loss": 21.1122996665928,
            "perplexity": math.inf,
            "num_examples": 500,
            "dataset": dataset,
            "file": str(root / "outputs" / "pythia-1.4b" / "seed_reps" / "seed_13" / "mia" / "utility_c5_teacher.json"),
        },
        {
            "seed": 17,
            "model": "c5_teacher",
            "mean_loss": 21.552093335229735,
            "perplexity": math.inf,
            "num_examples": 500,
            "dataset": dataset,
            "file": str(root / "outputs" / "pythia-1.4b" / "seed_reps" / "seed_17" / "mia" / "utility_c5_teacher.json"),
        },
        {
            "seed": 19,
            "model": "c5_teacher",
            "mean_loss": 31.713317265125525,
            "perplexity": math.inf,
            "num_examples": 500,
            "dataset": dataset,
            "file": str(root / "outputs" / "pythia-1.4b" / "seed_reps" / "seed_19" / "mia" / "utility_c5_teacher.json"),
        },
    ]
    return rows


def pythia_seed_c6_rows():
    mia_rows = []
    util_distill_rows = []
    util_holdout_rows = []
    for seed_dir in sorted(PYTHIA_SEED_ROOT.glob("seed_*")):
        try:
            seed = int(seed_dir.name.split("_", 1)[1])
        except Exception:
            continue
        mia_dir = seed_dir / "mia"
        for name in ["c6_student", "c6_teacher"]:
            path = mia_dir / f"{name}.json"
            if path.exists():
                data = read_json(path)
                mia_rows.append({
                    "seed": seed,
                    "condition": name,
                    "auroc": data["auroc"],
                    "auroc_doc": data["auroc_doc"],
                    "num_companies": len(data["per_company"]),
                    "file": str(path),
                })
        for name in ["utility_c6_student", "utility_c6_teacher"]:
            path = mia_dir / f"{name}.json"
            if path.exists():
                data = read_json(path)
                util_distill_rows.append({
                    "seed": seed,
                    "model": name.replace("utility_", ""),
                    "mean_loss": data["mean_loss"],
                    "perplexity": norm(data["perplexity"]),
                    "num_examples": data["num_examples"],
                    "dataset": data["dataset"],
                    "file": str(path),
                })
        for name in ["utility_c6_student_holdout", "utility_c6_teacher_holdout"]:
            path = mia_dir / f"{name}.json"
            if path.exists():
                data = read_json(path)
                util_holdout_rows.append({
                    "seed": seed,
                    "model": name.replace("utility_", ""),
                    "mean_loss": data["mean_loss"],
                    "perplexity": norm(data["perplexity"]),
                    "num_examples": data["num_examples"],
                    "dataset": data["dataset"],
                    "file": str(path),
                })
    return mia_rows, util_distill_rows, util_holdout_rows


def teacher_kl_rows():
    rows = []

    for scale, path in [
        ("1.4B", PYTHIA_ATTACK_DIR / "mia_teacher_attack" / "teacher_kl_targets_retains.json"),
        ("neo-1.3B", NEO_REPAIR_1P3B_DIR / "mia_teacher_attack" / "teacher_kl_targets_retains.json"),
    ]:
        if not path.exists():
            continue
        data = read_json(path)
        for ref_key, ref_data in data.items():
            for pool in ["target", "retain"]:
                entry = ref_data[pool]
                rows.append({
                    "scale": scale,
                    "reference": ref_key,
                    "pool": pool,
                    "kl_c6_to_ref_mean": entry["kl_c6_to_ref"]["mean"],
                    "kl_ref_to_c6_mean": entry["kl_ref_to_c6"]["mean"],
                    "js_proxy_mean": entry["js_proxy_mean"],
                    "num_tokens": entry["kl_c6_to_ref"]["n"],
                    "file": str(path),
                })
    return rows


def teacher_feature_attack_rows():
    rows = []
    for default_scale, path in [
        ("1.4B", PYTHIA_ATTACK_DIR / "mia_teacher_attack" / "teacher_feature_attack_summary.json"),
        ("neo-1.3B", NEO_REPAIR_1P3B_DIR / "mia_teacher_attack" / "teacher_feature_attack_summary.json"),
    ]:
        if not path.exists():
            continue
        data = read_json(path)
        for row in data.get("rows", []):
            if row.get("feature") == "num_tokens":
                continue
            rows.append({
                "scale": row.get("scale", default_scale),
                "reference": row.get("reference"),
                "candidate_pool": row.get("candidate_pool"),
                "positive_class": row.get("positive_class"),
                "negative_class": row.get("negative_class"),
                "feature": row.get("feature"),
                "auroc": row.get("auroc"),
                "ci_low": row.get("ci_low"),
                "ci_high": row.get("ci_high"),
                "positive_mean": row.get("positive_mean"),
                "negative_mean": row.get("negative_mean"),
                "n_positive": row.get("n_positive"),
                "n_negative": row.get("n_negative"),
                "file": row.get("file", str(path)),
            })
    return rows


def participation_threshold_rows():
    path = FOLLOWUP_DIR / "threshold_metrics.json"
    if not path.exists():
        return []
    data = read_json(path)
    rows = []
    for row in data.get("rows", []):
        metric = row.get("threshold_metrics", {})
        ba = metric.get("best_balanced_accuracy") or {}
        tpr_at = metric.get("tpr_at_fpr", {})
        topk = metric.get("topk", {})
        rows.append({
            "scale": row.get("scale"),
            "actor": row.get("actor"),
            "reference": row.get("reference"),
            "auroc": row.get("auroc"),
            "ci_low": (row.get("bootstrap_ci_95") or [None, None])[0],
            "ci_high": (row.get("bootstrap_ci_95") or [None, None])[1],
            "average_precision": metric.get("average_precision"),
            "tpr_at_1pct_fpr": (tpr_at.get("0.01") or {}).get("tpr"),
            "tpr_at_5pct_fpr": (tpr_at.get("0.05") or {}).get("tpr"),
            "tpr_at_10pct_fpr": (tpr_at.get("0.1") or {}).get("tpr"),
            "best_balanced_accuracy": ba.get("balanced_accuracy"),
            "precision_at_5": topk.get("precision_at_5"),
            "recall_at_5": topk.get("recall_at_5"),
            "file": row.get("file", str(path)),
        })
    return rows


def participation_gap_rows():
    path = FOLLOWUP_DIR / "student_teacher_gap.json"
    if not path.exists():
        return []
    data = read_json(path)
    rows = []
    for row in data.get("canonical", []):
        rows.append({
            "kind": "canonical",
            "scale": row.get("scale"),
            "reference": row.get("reference"),
            "student_minus_teacher_auroc": row.get("student_minus_teacher_auroc"),
            "ci_low": (row.get("bootstrap_ci_95") or [None, None])[0],
            "ci_high": (row.get("bootstrap_ci_95") or [None, None])[1],
            "p_value": (row.get("permutation_test") or {}).get("p_value_two_sided"),
            "sign_p_value": None,
            "n_positive": row.get("n_positive"),
            "n_negative": row.get("n_negative"),
            "num_seeds": None,
            "student_file": row.get("student_file"),
            "teacher_file": row.get("teacher_file"),
        })
    for row in data.get("seeded_pythia", []):
        rows.append({
            "kind": "seeded_pythia",
            "scale": row.get("scale"),
            "reference": row.get("reference"),
            "student_minus_teacher_auroc": row.get("mean_student_minus_teacher_auroc"),
            "ci_low": (row.get("bootstrap_ci_95") or [None, None])[0],
            "ci_high": (row.get("bootstrap_ci_95") or [None, None])[1],
            "p_value": None,
            "sign_p_value": (row.get("sign_test") or {}).get("p_value_two_sided"),
            "n_positive": None,
            "n_negative": None,
            "num_seeds": len(row.get("per_seed", [])),
            "student_file": None,
            "teacher_file": None,
        })
    return rows


def reference_mismatch_rows():
    path = FOLLOWUP_DIR / "reference_mismatch.json"
    if not path.exists():
        return []
    data = read_json(path)
    rows = []
    for row in data.get("rows", []):
        attack = row.get("attack", {})
        thresh = row.get("threshold_metrics", {})
        rows.append({
            "scale": row.get("scale"),
            "c6_seed": row.get("c6_seed"),
            "reference_kind": row.get("reference_kind"),
            "reference_seed": row.get("reference_seed"),
            "reference_condition": row.get("reference_condition"),
            "auroc": attack.get("auroc"),
            "ci_low": (attack.get("bootstrap_ci_95") or [None, None])[0],
            "ci_high": (attack.get("bootstrap_ci_95") or [None, None])[1],
            "tpr_at_5pct_fpr": ((thresh.get("tpr_at_fpr") or {}).get("0.05") or {}).get("tpr"),
            "best_balanced_accuracy": ((thresh.get("best_balanced_accuracy") or {}).get("balanced_accuracy")),
            "file": str(path),
        })
    return rows


def placebo_rows():
    rows = []
    for scale, base_dir, ref_map in [
        ("1.4B", PYTHIA_PLACEBO_DIR, [("C3", "mia_c6_placebo_deletion_attack_target_vs_retain.json"), ("C1", "mia_c6_placebo_deletion_attack_target_vs_retain_c1ref.json")]),
        ("neo-1.3B", NEO_PLACEBO_1P3B_DIR, [("C3", "mia_c6_placebo_deletion_attack_target_vs_retain.json"), ("C1", "mia_c6_placebo_deletion_attack_target_vs_retain_c1ref.json")]),
    ]:
        for reference, filename in ref_map:
            path = base_dir / filename
            if not path.exists():
                continue
            data = read_json(path)
            ci = data.get("bootstrap_ci_95") or [None, None]
            rows.append({
                "status": "completed",
                "reason": f"{scale} placebo random-forget NPO retain-pool attack ref={reference}",
                "auroc": data.get("auroc"),
                "ci_low": ci[0],
                "ci_high": ci[1],
                "file": str(path),
            })
    if rows:
        return rows
    path = FOLLOWUP_DIR / "placebo_control.json"
    if not path.exists():
        return []
    data = read_json(path)
    return [{
        "status": data.get("status"),
        "reason": data.get("reason"),
        "auroc": data.get("auroc"),
        "ci_low": (data.get("bootstrap_ci_95") or [None, None])[0] if data.get("bootstrap_ci_95") else None,
        "ci_high": (data.get("bootstrap_ci_95") or [None, None])[1] if data.get("bootstrap_ci_95") else None,
        "file": data.get("file", str(path)),
    }]


def placebo_attack_rows():
    rows = []
    for scale, base_dir, attack_map in [
        ("1.4B", PYTHIA_PLACEBO_DIR, [("mia_c6_placebo_deletion_attack_target_vs_retain.json", "C6_placebo_del_retain_refC3", "C6_placebo_minus_C3"), ("mia_c6_placebo_deletion_attack_target_vs_retain_c1ref.json", "C6_placebo_del_retain_refC1", "C6_placebo_minus_C1")]),
        ("neo-1.3B", NEO_PLACEBO_1P3B_DIR, [("mia_c6_placebo_deletion_attack_target_vs_retain.json", "C6_placebo_del_retain_refC3", "C6_placebo_minus_C3"), ("mia_c6_placebo_deletion_attack_target_vs_retain_c1ref.json", "C6_placebo_del_retain_refC1", "C6_placebo_minus_C1")]),
    ]:
        for filename, attack_id, reference in attack_map:
            path = base_dir / filename
            if not path.exists():
                continue
            data = read_json(path)
            ci = data.get("bootstrap_ci_95", [None, None])
            rows.append({
                "scale": scale,
                "attack_id": attack_id,
                "reference": reference,
                "candidate_pool": "retained_companies",
                "positive_class": data.get("positive_class"),
                "negative_class": data.get("negative_class"),
                "score_definition": "delta_mean_loss",
                "auroc": data.get("auroc"),
                "ci_low": ci[0],
                "ci_high": ci[1],
                "positive_mean": data.get("target_mean_delta"),
                "negative_mean": data.get("retain_mean_delta"),
                "n_positive": data.get("num_targets"),
                "n_negative": data.get("num_retained"),
                "file": str(path),
            })
    return rows


def pythia_c6_mia_rows():
    rows = []
    for label, path in [
        ("C6 teacher", PYTHIA_MIA_DIR / "c6_teacher.json"),
        ("C6 student", PYTHIA_MIA_DIR / "c6_student.json"),
    ]:
        if not path.exists():
            continue
        data = read_json(path)
        rows.append({
            "scale": "1.4B",
            "condition": label,
            "auroc": data["auroc"],
            "num_companies": len(data["per_company"]),
            "status": "ok",
            "file": str(path),
        })
    return rows


def seeded_placebo_attack_rows():
    rows = []
    configs = [
        ("1.4B", PYTHIA_SEEDED_PLACEBO_DIR),
        ("neo-1.3B", NEO_SEEDED_PLACEBO_DIR),
    ]
    for scale, base_dir in configs:
        if not base_dir.exists():
            continue
        for path in sorted(base_dir.glob("seed_reps/seed_*/placebo_c6_c1_retain/mia_c6_placebo_deletion_attack_target_vs_retain_c1ref.json")):
            data = read_json(path)
            ci = data.get("bootstrap_ci_95", [None, None])
            seed = data.get("seed")
            if seed is None:
                try:
                    seed = int(path.parts[path.parts.index("seed_reps") + 1].split("_")[-1])
                except Exception:
                    seed = None
            rows.append({
                "scale": scale,
                "seed": seed,
                "attack_id": "C6_placebo_del_retain_refC1_seeded",
                "reference": "C1",
                "candidate_pool": "retained_companies",
                "positive_class": data.get("positive_class"),
                "negative_class": data.get("negative_class"),
                "score_definition": "delta_mean_loss",
                "auroc": data.get("auroc"),
                "ci_low": ci[0],
                "ci_high": ci[1],
                "positive_mean": data.get("target_mean_delta"),
                "negative_mean": data.get("retain_mean_delta"),
                "n_positive": data.get("num_targets"),
                "n_negative": data.get("num_retained"),
                "file": str(path),
            })
    return rows


def seeded_placebo_summary_rows():
    rows = []
    configs = [
        ("1.4B", PYTHIA_SEEDED_PLACEBO_DIR),
        ("neo-1.3B", NEO_SEEDED_PLACEBO_DIR),
    ]
    for scale, base_dir in configs:
        path = base_dir / "seed_c6_placebo_deletion_attack_target_vs_retain_c1ref_summary.json"
        if not path.exists():
            continue
        data = read_json(path)
        rows.append({
            "scale": scale,
            "attack_id": "C6_placebo_del_retain_refC1_seeded",
            "reference": data.get("reference", "C1"),
            "candidate_pool": data.get("negative_class", "retained_companies"),
            "result_scope": "seeded placebo C1-reference retain-pool",
            "mean_auroc": data.get("mean_auroc"),
            "std_auroc": data.get("sd_auroc"),
            "n_seeds": data.get("n_runs"),
            "file": str(path),
        })
    return rows


def pythia_c6_utility_rows(kind):
    suffix = "" if kind == "distill" else "_holdout"
    rows = []
    for model in ["c6_student", "c6_teacher"]:
        path = PYTHIA_MIA_DIR / f"utility_{model}{suffix}.json"
        if not path.exists():
            continue
        data = read_json(path)
        rows.append({
            "scale": "1.4B",
            "model": f"{model}{'_holdout' if kind == 'holdout' else ''}",
            "mean_loss": data["mean_loss"],
            "perplexity": norm(data["perplexity"]),
            "num_examples": data["num_examples"],
            "dataset": data["dataset"],
            "file": str(path),
        })
    return rows


def pythia_c6_paired_rows():
    def load_mean_loss_map(path: Path):
        data = read_json(path)
        return {
            cik: values["mean_loss"]
            for cik, values in data["per_company"].items()
            if values.get("mean_loss") is not None
        }

    paths = {
        "C1": PYTHIA_MIA_DIR / "c1_student.json",
        "C2": PYTHIA_MIA_DIR / "c2_student.json",
        "C3": PYTHIA_MIA_DIR / "c3_student.json",
        "C6": PYTHIA_MIA_DIR / "c6_student.json",
    }
    if not all(path.exists() for path in paths.values()):
        return []

    loss_maps = {label: load_mean_loss_map(path) for label, path in paths.items()}

    def build_row(label_a: str, label_b: str):
        common = sorted(set(loss_maps[label_a]) & set(loss_maps[label_b]))
        if not common:
            return None
        a_vals = [loss_maps[label_a][k] for k in common]
        b_vals = [loss_maps[label_b][k] for k in common]
        diffs = [a - b for a, b in zip(a_vals, b_vals)]
        sign = sign_test(a_vals, b_vals)
        perm = paired_permutation_test(a_vals, b_vals, n=20000, seed=13)
        ci_low, ci_high = bootstrap_ci_diff(a_vals, b_vals, n=2000, seed=13)
        return {
            "scale": "1.4B",
            "comparison": f"{label_a}_vs_{label_b}",
            "mean_diff": sum(diffs) / len(diffs),
            "ci_low": ci_low,
            "ci_high": ci_high,
            "sign_p": sign["p_value_two_sided"],
            "sign_pos": sign["pos"],
            "sign_neg": sign["neg"],
            "perm_p": perm["p_value_two_sided"],
            "n": len(common),
        }

    rows = []
    for pair in [("C6", "C1"), ("C6", "C2"), ("C6", "C3")]:
        row = build_row(*pair)
        if row is not None:
            rows.append(row)
    return rows


def pythia_c6_attack_rows():
    rows = []

    path = PYTHIA_ATTACK_DIR / "mia_c6_deletion_attack.json"
    if path.exists():
        data = read_json(path)
        mapping = [
            (
                "C6_del_nonmember_canonical_refA",
                "C6_minus_C3",
                "nonmembers_canonical",
                "deleted_targets",
                "nonmembers",
                "delta_mean_loss",
                "reference_A_canonical_C6_minus_C3",
                "target_delta_mean",
                "nonmember_delta_mean",
                "targets_n",
                "nonmembers_n",
                "auroc",
            ),
            (
                "C6_del_nonmember_matched_refA",
                "C6_minus_C3",
                "nonmembers_matched",
                "deleted_targets",
                "matched_nonmembers",
                "delta_mean_loss",
                "reference_A_matched_C6_minus_C3",
                "target_delta_mean",
                "nonmember_delta_mean",
                "targets_n",
                "nonmembers_n",
                "auroc",
            ),
            (
                "C6_del_nonmember_canonical_refB",
                "C6_absolute_loss",
                "nonmembers_canonical",
                "deleted_targets",
                "nonmembers",
                "absolute_mean_loss",
                "reference_B_absolute_loss_canonical",
                None,
                None,
                None,
                None,
                "auroc_if_score_is_high_loss_for_deletion",
            ),
            (
                "C6_del_nonmember_matched_refB",
                "C6_absolute_loss",
                "nonmembers_matched",
                "deleted_targets",
                "matched_nonmembers",
                "absolute_mean_loss",
                "reference_B_absolute_loss_matched",
                None,
                None,
                None,
                None,
                "auroc_if_score_is_high_loss_for_deletion",
            ),
        ]
        for attack_id, reference, pool, pos_cls, neg_cls, score_def, key, pos_mean_key, neg_mean_key, n_pos_key, n_neg_key, auroc_key in mapping:
            if key not in data:
                continue
            entry = data[key]
            rows.append({
                "scale": "1.4B",
                "attack_id": attack_id,
                "reference": reference,
                "candidate_pool": pool,
                "positive_class": pos_cls,
                "negative_class": neg_cls,
                "score_definition": score_def,
                "auroc": entry.get(auroc_key),
                "ci_low": None,
                "ci_high": None,
                "positive_mean": entry.get(pos_mean_key) if pos_mean_key else None,
                "negative_mean": entry.get(neg_mean_key) if neg_mean_key else None,
                "n_positive": entry.get(n_pos_key) if n_pos_key else None,
                "n_negative": entry.get(n_neg_key) if n_neg_key else None,
                "file": str(path),
            })

    for filename, attack_id, reference in [
        ("mia_c6_deletion_attack_target_vs_retain.json", "C6_del_retain_refC3", "C6_minus_C3"),
        ("mia_c6_deletion_attack_target_vs_retain_c1ref.json", "C6_del_retain_refC1", "C6_minus_C1"),
    ]:
        path = PYTHIA_ATTACK_DIR / filename
        if not path.exists():
            continue
        data = read_json(path)
        ci = data.get("bootstrap_ci_95", [None, None])
        rows.append({
            "scale": "1.4B",
            "attack_id": attack_id,
            "reference": reference,
            "candidate_pool": "retained_companies",
            "positive_class": data.get("positive_class"),
            "negative_class": data.get("negative_class"),
            "score_definition": "delta_mean_loss",
            "auroc": data.get("auroc"),
            "ci_low": ci[0],
            "ci_high": ci[1],
            "positive_mean": data.get("target_mean_delta"),
            "negative_mean": data.get("retain_mean_delta"),
            "n_positive": data.get("num_targets"),
            "n_negative": data.get("num_retained"),
            "file": str(path),
        })

    for filename, attack_id, reference in [
        ("mia_c6_teacher_deletion_attack_target_vs_retain_c3ref.json", "C6_teacher_del_retain_refC3", "C6_teacher_minus_C3_teacher"),
        ("mia_c6_teacher_deletion_attack_target_vs_retain_c1ref.json", "C6_teacher_del_retain_refC1", "C6_teacher_minus_C1_teacher"),
    ]:
        path = PYTHIA_ATTACK_DIR / filename
        if not path.exists():
            continue
        data = read_json(path)
        ci = data.get("bootstrap_ci_95", [None, None])
        rows.append({
            "scale": "1.4B",
            "attack_id": attack_id,
            "reference": reference,
            "candidate_pool": "retained_companies",
            "positive_class": data.get("positive_class"),
            "negative_class": data.get("negative_class"),
            "score_definition": "delta_mean_loss",
            "auroc": data.get("auroc"),
            "ci_low": ci[0],
            "ci_high": ci[1],
            "positive_mean": data.get("target_mean_delta"),
            "negative_mean": data.get("retain_mean_delta"),
            "n_positive": data.get("num_targets"),
            "n_negative": data.get("num_retained"),
            "file": str(path),
        })

    return rows


def pythia_seed_c6_attack_rows():
    rows = []
    summary = PYTHIA_ATTACK_DIR / "seed_c6_deletion_attack_summary.json"
    if summary.exists():
        data = read_json(summary)
        for row in data.get("rows", []):
            key = row.get("key")
            if key == "reference_A_canonical_C6_minus_C1":
                attack_id = "C6_del_seed_nonmember_refC1"
                score_definition = "delta_mean_loss"
            elif key == "reference_A_canonical_C6_minus_C3":
                attack_id = "C6_del_seed_nonmember_refC3"
                score_definition = "delta_mean_loss"
            elif key == "reference_B_absolute_loss_canonical":
                attack_id = "C6_del_seed_nonmember_refB"
                score_definition = "absolute_mean_loss"
            else:
                continue
            rows.append({
                "seed": row.get("seed"),
                "attack_id": attack_id,
                "reference": row.get("reference"),
                "candidate_pool": "nonmembers_canonical",
                "positive_class": "deleted_targets",
                "negative_class": "nonmembers",
                "score_definition": score_definition,
                "auroc": row.get("auroc"),
                "ci_low": row.get("ci_low"),
                "ci_high": row.get("ci_high"),
                "positive_mean": row.get("positive_mean"),
                "negative_mean": row.get("negative_mean"),
                "n_positive": row.get("n_positive"),
                "n_negative": row.get("n_negative"),
                "file": row.get("file"),
            })

    retain_summary = PYTHIA_ATTACK_DIR / "seed_c6_deletion_attack_target_vs_retain_summary.json"
    if retain_summary.exists():
        data = read_json(retain_summary)
        for row in data.get("rows", []):
            attack_id = row.get("attack_id")
            if attack_id not in {"C6_del_seed_retain_refC1", "C6_del_seed_retain_refC3"}:
                continue
            rows.append({
                "seed": row.get("seed"),
                "attack_id": attack_id,
                "reference": row.get("reference"),
                "candidate_pool": "retained_companies",
                "positive_class": "deleted_targets",
                "negative_class": "retained_companies",
                "score_definition": "delta_mean_loss",
                "auroc": row.get("auroc"),
                "ci_low": row.get("ci_low"),
                "ci_high": row.get("ci_high"),
                "positive_mean": row.get("positive_mean"),
                "negative_mean": row.get("negative_mean"),
                "n_positive": row.get("n_positive"),
                "n_negative": row.get("n_negative"),
                "file": row.get("file"),
            })
    return rows


def pythia_seed_c6_attack_summary_rows():
    rows = []
    summary = PYTHIA_ATTACK_DIR / "seed_c6_deletion_attack_summary.json"
    if summary.exists():
        data = read_json(summary)
        for key, attack_id, reference in [
            ("reference_A_canonical_C6_minus_C1", "C6_del_seed_nonmember_refC1", "C6_minus_C1"),
            ("reference_A_canonical_C6_minus_C3", "C6_del_seed_nonmember_refC3", "C6_minus_C3"),
            ("reference_B_absolute_loss_canonical", "C6_del_seed_nonmember_refB", "C6_absolute_loss"),
        ]:
            if key not in data:
                continue
            entry = data[key]
            rows.append({
                "attack_id": attack_id,
                "reference": reference,
                "candidate_pool": "nonmembers_canonical",
                "result_scope": "seed stability only; not headline retain-pool attack",
                "mean_auroc": entry.get("mean_auroc"),
                "std_auroc": entry.get("std_auroc"),
                "n_seeds": entry.get("n_seeds"),
                "file": str(summary),
            })

    retain_summary = PYTHIA_ATTACK_DIR / "seed_c6_deletion_attack_target_vs_retain_summary.json"
    if retain_summary.exists():
        data = read_json(retain_summary)
        for attack_id, reference in [
            ("C6_del_seed_retain_refC1", "C6_minus_C1"),
            ("C6_del_seed_retain_refC3", "C6_minus_C3"),
        ]:
            if attack_id not in data:
                continue
            entry = data[attack_id]
            rows.append({
                "attack_id": attack_id,
                "reference": reference,
                "candidate_pool": "retained_companies",
                "result_scope": "headline retain-pool seed replication",
                "mean_auroc": entry.get("mean_auroc"),
                "std_auroc": entry.get("std_auroc"),
                "n_seeds": entry.get("n_seeds"),
                "file": str(retain_summary),
            })
    return rows


def pythia_seed_c6_teacher_attack_rows():
    rows = []
    summary = PYTHIA_ATTACK_DIR / "seed_c6_teacher_deletion_attack_target_vs_retain_summary.json"
    if not summary.exists():
        return rows

    data = read_json(summary)
    for row in data.get("rows", []):
        attack_id = row.get("attack_id")
        if attack_id not in {"C6_teacher_del_seed_retain_refC1", "C6_teacher_del_seed_retain_refC3"}:
            continue
        rows.append({
            "seed": row.get("seed"),
            "attack_id": attack_id,
            "reference": row.get("reference"),
            "candidate_pool": "retained_companies",
            "positive_class": "deleted_targets",
            "negative_class": "retained_companies",
            "score_definition": "delta_mean_loss",
            "auroc": row.get("auroc"),
            "ci_low": row.get("ci_low"),
            "ci_high": row.get("ci_high"),
            "positive_mean": row.get("positive_mean"),
            "negative_mean": row.get("negative_mean"),
            "n_positive": row.get("n_positive"),
            "n_negative": row.get("n_negative"),
            "file": row.get("file"),
        })
    return rows


def pythia_seed_c6_teacher_attack_summary_rows():
    rows = []
    summary = PYTHIA_ATTACK_DIR / "seed_c6_teacher_deletion_attack_target_vs_retain_summary.json"
    if not summary.exists():
        return rows

    data = read_json(summary)
    for attack_id, reference in [
        ("C6_teacher_del_seed_retain_refC1", "C6_teacher_minus_C1_teacher"),
        ("C6_teacher_del_seed_retain_refC3", "C6_teacher_minus_C3_teacher"),
    ]:
        if attack_id not in data:
            continue
        entry = data[attack_id]
        rows.append({
            "attack_id": attack_id,
            "reference": reference,
            "candidate_pool": "retained_companies",
            "result_scope": "teacher-side headline retain-pool seed replication",
            "mean_auroc": entry.get("mean_auroc"),
            "std_auroc": entry.get("std_auroc"),
            "n_seeds": entry.get("n_seeds"),
            "file": str(summary),
        })
    return rows


def neo_c6_attack_rows():
    rows = []
    path = NEO_REPAIR_1P3B_DIR / "mia_c6_deletion_attack.json"
    if not path.exists():
        return rows

    data = read_json(path)
    mapping = [
        (
            "C6_del_nonmember_canonical_refC1",
            "C6_minus_C1",
            "reference_A_canonical_C6_minus_C1",
            "target_delta_mean",
            "nonmember_delta_mean",
            "targets_n",
            "nonmembers_n",
            "auroc",
        ),
        (
            "C6_del_nonmember_canonical_refC3",
            "C6_minus_C3",
            "reference_A_canonical_C6_minus_C3",
            "target_delta_mean",
            "nonmember_delta_mean",
            "targets_n",
            "nonmembers_n",
            "auroc",
        ),
        (
            "C6_del_nonmember_canonical_refB",
            "C6_absolute_loss",
            "reference_B_absolute_loss_canonical",
            None,
            None,
            None,
            None,
            "auroc_if_score_is_high_loss_for_deletion",
        ),
    ]
    for attack_id, reference, key, pos_mean_key, neg_mean_key, n_pos_key, n_neg_key, auroc_key in mapping:
        if key not in data:
            continue
        entry = data[key]
        rows.append({
            "scale": "neo-1.3B",
            "attack_id": attack_id,
            "reference": reference,
            "candidate_pool": "nonmembers_canonical",
            "positive_class": "deleted_targets",
            "negative_class": "nonmembers",
            "score_definition": "absolute_mean_loss" if "absolute" in reference else "delta_mean_loss",
            "auroc": entry.get(auroc_key),
            "ci_low": None,
            "ci_high": None,
            "positive_mean": entry.get(pos_mean_key) if pos_mean_key else None,
            "negative_mean": entry.get(neg_mean_key) if neg_mean_key else None,
            "n_positive": entry.get(n_pos_key) if n_pos_key else None,
            "n_negative": entry.get(n_neg_key) if n_neg_key else None,
            "file": str(path),
        })

    for filename, attack_id, reference in [
        ("mia_c6_deletion_attack_target_vs_retain.json", "C6_del_retain_refC3", "C6_minus_C3"),
        ("mia_c6_deletion_attack_target_vs_retain_c1ref.json", "C6_del_retain_refC1", "C6_minus_C1"),
    ]:
        retain_path = NEO_REPAIR_1P3B_DIR / filename
        if not retain_path.exists():
            continue
        entry = read_json(retain_path)
        ci = entry.get("bootstrap_ci_95", [None, None])
        rows.append({
            "scale": "neo-1.3B",
            "attack_id": attack_id,
            "reference": reference,
            "candidate_pool": "retained_companies",
            "positive_class": entry.get("positive_class"),
            "negative_class": entry.get("negative_class"),
            "score_definition": "delta_mean_loss",
            "auroc": entry.get("auroc"),
            "ci_low": ci[0],
            "ci_high": ci[1],
            "positive_mean": entry.get("target_mean_delta"),
            "negative_mean": entry.get("retain_mean_delta"),
            "n_positive": entry.get("num_targets"),
            "n_negative": entry.get("num_retained"),
            "file": str(retain_path),
        })
    return rows


def neo_teacher_c6_attack_rows():
    path = NEO_REPAIR_1P3B_DIR / "mia_c6_teacher_deletion_attack.json"
    data = read_json(path) if path.exists() else {}
    mapping = [
        (
            "C6_teacher_del_nonmember_refC1",
            "C6_teacher_minus_C1_teacher",
            "delta_mean_loss",
            "reference_A_canonical_C6_teacher_minus_C1_teacher",
            "target_delta_mean",
            "nonmember_delta_mean",
            "targets_n",
            "nonmembers_n",
            "auroc",
        ),
        (
            "C6_teacher_del_nonmember_refB",
            "C6_teacher_absolute_loss",
            "absolute_mean_loss",
            "reference_B_absolute_C6_teacher_loss_canonical",
            "target_loss_mean",
            "nonmember_loss_mean",
            "targets_n",
            "nonmembers_n",
            "auroc_if_score_is_high_loss_for_deletion",
        ),
    ]
    rows = []
    for attack_id, reference, score_definition, key, pos_mean_key, neg_mean_key, n_pos_key, n_neg_key, auroc_key in mapping:
        if key not in data:
            continue
        entry = data[key]
        ci = entry.get("bootstrap_ci_95", [None, None])
        rows.append({
            "scale": "neo-1.3B",
            "attack_id": attack_id,
            "reference": reference,
            "candidate_pool": "nonmembers_canonical",
            "positive_class": "deleted_targets",
            "negative_class": "nonmembers",
            "score_definition": score_definition,
            "auroc": entry.get(auroc_key),
            "ci_low": ci[0],
            "ci_high": ci[1],
            "positive_mean": entry.get(pos_mean_key),
            "negative_mean": entry.get(neg_mean_key),
            "n_positive": entry.get(n_pos_key),
            "n_negative": entry.get(n_neg_key),
            "file": str(path),
        })
    for filename, attack_id, reference in [
        ("mia_c6_teacher_deletion_attack_target_vs_retain_c3ref.json", "C6_teacher_del_retain_refC3", "C6_teacher_minus_C3_teacher"),
        ("mia_c6_teacher_deletion_attack_target_vs_retain_c1ref.json", "C6_teacher_del_retain_refC1", "C6_teacher_minus_C1_teacher"),
    ]:
        retain_path = NEO_REPAIR_1P3B_DIR / filename
        if not retain_path.exists():
            continue
        retain = read_json(retain_path)
        ci = retain.get("bootstrap_ci_95", [None, None])
        rows.append({
            "scale": "neo-1.3B",
            "attack_id": attack_id,
            "reference": reference,
            "candidate_pool": "retained_companies",
            "positive_class": retain.get("positive_class"),
            "negative_class": retain.get("negative_class"),
            "score_definition": "delta_mean_loss",
            "auroc": retain.get("auroc"),
            "ci_low": ci[0],
            "ci_high": ci[1],
            "positive_mean": retain.get("target_mean_delta"),
            "negative_mean": retain.get("retain_mean_delta"),
            "n_positive": retain.get("num_targets"),
            "n_negative": retain.get("num_retained"),
            "file": str(retain_path),
        })
    return rows


def neo_teacher_deletion_bound_rows():
    path = NEO_REPAIR_1P3B_DIR / "mia_c6_teacher_deletion_attack.json"
    if not path.exists():
        return []
    data = read_json(path)
    rows = []
    for bound_id, entry in data.get("explicit_bounds", {}).items():
        if bound_id == "C6_teacher_minus_C3_teacher" and (NEO_REPAIR_1P3B_DIR / "mia_c6_teacher_deletion_attack_target_vs_retain_c3ref.json").exists():
            continue
        if bound_id == "retain_pool_C6_teacher_minus_C1_teacher" and (NEO_REPAIR_1P3B_DIR / "mia_c6_teacher_deletion_attack_target_vs_retain_c1ref.json").exists():
            continue
        rows.append({
            "scale": "neo-1.3B",
            "bound_id": bound_id,
            "auroc_lower_bound": entry.get("auroc_lower_bound"),
            "auroc_upper_bound": entry.get("auroc_upper_bound"),
            "reason": entry.get("reason"),
            "required_missing_files": ", ".join(entry.get("required_missing_files", [])),
            "file": str(path),
        })
    return rows


def main():
    wb = load_workbook(WORKBOOK)
    rename_sheet_if_present(wb, "seed_teacher_deletion_attacks_1.4B", SEED_TEACHER_ATTACK_SHEET)
    rename_sheet_if_present(wb, "seed_teacher_deletion_summary_1.4B", SEED_TEACHER_SUMMARY_SHEET)
    (
        neo_2p7b_mia_rows,
        neo_2p7b_paired_rows,
        neo_2p7b_matched_rows,
        neo_2p7b_paired_matched_rows,
        neo_2p7b_utility_distill_rows,
        neo_2p7b_utility_holdout_rows,
    ) = neo_2p7b_manual_rows()
    (
        neo_c6_mia_rows,
        neo_c6_utility_distill_rows,
        neo_c6_utility_holdout_rows,
        neo_c6_run_metrics_rows,
    ) = neo_c6_manual_rows()
    (
        neo_seed_c6_mia_rows,
        neo_seed_c6_utility_distill_rows,
        neo_seed_c6_utility_holdout_rows,
        neo_seed_c6_run_metrics_rows,
    ) = neo_seed_c6_manual_rows()

    mia_ws = ensure_sheet(wb, "mia_auroc", ["scale", "condition", "auroc", "num_companies", "status", "file"])
    for row in direct_mia_rows():
        upsert(mia_ws, ["scale", "condition"], row)
    for row in pythia_c6_mia_rows():
        upsert(mia_ws, ["scale", "condition"], row)
    for row in neo_2p7b_mia_rows:
        upsert(mia_ws, ["scale", "condition"], row)
    for row in neo_c6_mia_rows:
        upsert(mia_ws, ["scale", "condition"], row)
    sort_sheet(mia_ws, ["scale", "condition"])

    paired_ws = ensure_sheet(wb, "paired_stats", ["scale", "comparison", "mean_diff", "ci_low", "ci_high", "sign_p", "sign_pos", "sign_neg", "perm_p", "n"])
    for row in paired_rows():
        upsert(paired_ws, ["scale", "comparison"], row)
    for row in pythia_c6_paired_rows():
        upsert(paired_ws, ["scale", "comparison"], row)
    for row in neo_2p7b_paired_rows:
        upsert(paired_ws, ["scale", "comparison"], row)
    sort_sheet(paired_ws, ["scale", "comparison"])

    util_distill_ws = ensure_sheet(wb, "utility_distill", ["scale", "model", "mean_loss", "perplexity", "num_examples", "dataset", "file"])
    for row in utility_rows("distill"):
        upsert(util_distill_ws, ["scale", "model"], row)
    for row in pythia_c6_utility_rows("distill"):
        upsert(util_distill_ws, ["scale", "model"], row)
    for row in neo_2p7b_utility_distill_rows:
        upsert(util_distill_ws, ["scale", "model"], row)
    for row in neo_c6_utility_distill_rows:
        upsert(util_distill_ws, ["scale", "model"], row)
    sort_sheet(util_distill_ws, ["scale", "model"])

    util_holdout_ws = ensure_sheet(wb, "utility_holdout", ["scale", "model", "mean_loss", "perplexity", "num_examples", "dataset", "file"])
    for row in utility_rows("holdout"):
        upsert(util_holdout_ws, ["scale", "model"], row)
    for row in pythia_c6_utility_rows("holdout"):
        upsert(util_holdout_ws, ["scale", "model"], row)
    for row in neo_2p7b_utility_holdout_rows:
        upsert(util_holdout_ws, ["scale", "model"], row)
    for row in neo_c6_utility_holdout_rows:
        upsert(util_holdout_ws, ["scale", "model"], row)
    sort_sheet(util_holdout_ws, ["scale", "model"])

    run_metrics_ws = ensure_sheet(
        wb,
        "run_metrics",
        [
            "scale",
            "condition",
            "seed",
            "method",
            "total_steps",
            "initial_forget_loss",
            "final_forget_loss",
            "initial_retain_loss",
            "final_retain_loss",
            "final_forget_log_ratio",
            "ref_checksum_match",
            "file",
        ],
    )
    for row in neo_c6_run_metrics_rows:
        upsert(run_metrics_ws, ["scale", "condition", "seed"], row)
    sort_sheet(run_metrics_ws, ["scale", "condition", "seed"])

    matched_ws = ensure_sheet(wb, "mia_matched_neo_1.3B", ["condition", "auroc", "auroc_ci95_low", "auroc_ci95_high", "auroc_doc", "mean_loss_ratio", "loss_ratio_ci95_low", "loss_ratio_ci95_high", "n_companies", "n_nonmember_companies"])
    matched_rows_data, paired_matched_rows = matched_rows()
    for row in matched_rows_data:
        upsert(matched_ws, ["condition"], row)
    sort_sheet(matched_ws, ["condition"])

    matched_2p7b_ws = ensure_sheet(wb, "mia_matched_neo_2.7B", ["condition", "auroc", "auroc_ci95_low", "auroc_ci95_high", "auroc_doc", "mean_loss_ratio", "loss_ratio_ci95_low", "loss_ratio_ci95_high", "n_companies", "n_nonmember_companies"])
    for row in neo_2p7b_matched_rows:
        upsert(matched_2p7b_ws, ["condition"], row)
    sort_sheet(matched_2p7b_ws, ["condition"])

    paired_matched_ws = ensure_sheet(wb, "paired_matched", ["scale", "comparison", "mean_diff", "ci_low", "ci_high", "sign_p", "sign_pos", "sign_neg", "perm_p", "n"])
    for row in paired_matched_rows:
        upsert(paired_matched_ws, ["scale", "comparison"], row)
    for row in neo_2p7b_paired_matched_rows:
        upsert(paired_matched_ws, ["scale", "comparison"], row)
    sort_sheet(paired_matched_ws, ["scale", "comparison"])

    attack_ws = ensure_sheet(
        wb,
        "deletion_attacks",
        ["scale", "attack_id", "reference", "candidate_pool", "positive_class", "negative_class", "score_definition", "auroc", "ci_low", "ci_high", "positive_mean", "negative_mean", "n_positive", "n_negative", "file"],
    )
    for row in pythia_c6_attack_rows():
        upsert(attack_ws, ["scale", "attack_id"], row)
    for row in neo_c6_attack_rows():
        upsert(attack_ws, ["scale", "attack_id"], row)
    for row in neo_teacher_c6_attack_rows():
        upsert(attack_ws, ["scale", "attack_id"], row)
    for row in placebo_attack_rows():
        upsert(attack_ws, ["scale", "attack_id"], row)
    sort_sheet(attack_ws, ["scale", "attack_id"])

    seed_attack_ws = ensure_sheet(
        wb,
        "seed_deletion_attacks_1.4B",
        ["seed", "attack_id", "reference", "candidate_pool", "positive_class", "negative_class", "score_definition", "auroc", "ci_low", "ci_high", "positive_mean", "negative_mean", "n_positive", "n_negative", "file"],
    )
    for row in pythia_seed_c6_attack_rows():
        upsert(seed_attack_ws, ["seed", "attack_id"], row)
    sort_sheet(seed_attack_ws, ["seed", "attack_id"])

    seed_attack_summary_ws = ensure_sheet(
        wb,
        "seed_deletion_summary_1.4B",
        ["attack_id", "reference", "candidate_pool", "result_scope", "mean_auroc", "std_auroc", "n_seeds", "file"],
    )
    for row in pythia_seed_c6_attack_summary_rows():
        upsert(seed_attack_summary_ws, ["attack_id"], row)
    sort_sheet(seed_attack_summary_ws, ["attack_id"])

    seed_teacher_attack_ws = ensure_sheet(
        wb,
        SEED_TEACHER_ATTACK_SHEET,
        ["seed", "attack_id", "reference", "candidate_pool", "positive_class", "negative_class", "score_definition", "auroc", "ci_low", "ci_high", "positive_mean", "negative_mean", "n_positive", "n_negative", "file"],
    )
    for row in pythia_seed_c6_teacher_attack_rows():
        upsert(seed_teacher_attack_ws, ["seed", "attack_id"], row)
    sort_sheet(seed_teacher_attack_ws, ["seed", "attack_id"])

    seed_teacher_attack_summary_ws = ensure_sheet(
        wb,
        SEED_TEACHER_SUMMARY_SHEET,
        ["attack_id", "reference", "candidate_pool", "result_scope", "mean_auroc", "std_auroc", "n_seeds", "file"],
    )
    for row in pythia_seed_c6_teacher_attack_summary_rows():
        upsert(seed_teacher_attack_summary_ws, ["attack_id"], row)
    sort_sheet(seed_teacher_attack_summary_ws, ["attack_id"])

    seed_placebo_attack_ws = ensure_sheet(
        wb,
        "seed_placebo_attacks",
        ["scale", "seed", "attack_id", "reference", "candidate_pool", "positive_class", "negative_class", "score_definition", "auroc", "ci_low", "ci_high", "positive_mean", "negative_mean", "n_positive", "n_negative", "file"],
    )
    for row in seeded_placebo_attack_rows():
        upsert(seed_placebo_attack_ws, ["scale", "seed", "attack_id"], row)
    sort_sheet(seed_placebo_attack_ws, ["scale", "seed", "attack_id"])

    seed_placebo_summary_ws = ensure_sheet(
        wb,
        "seed_placebo_summary",
        ["scale", "attack_id", "reference", "candidate_pool", "result_scope", "mean_auroc", "std_auroc", "n_seeds", "file"],
    )
    for row in seeded_placebo_summary_rows():
        upsert(seed_placebo_summary_ws, ["scale", "attack_id"], row)
    sort_sheet(seed_placebo_summary_ws, ["scale", "attack_id"])

    teacher_bound_ws = ensure_sheet(
        wb,
        "teacher_deletion_bounds",
        ["scale", "bound_id", "auroc_lower_bound", "auroc_upper_bound", "reason", "required_missing_files", "file"],
    )
    neo_teacher_retain_c1_path = NEO_REPAIR_1P3B_DIR / "mia_c6_teacher_deletion_attack_target_vs_retain_c1ref.json"
    neo_teacher_retain_c3_path = NEO_REPAIR_1P3B_DIR / "mia_c6_teacher_deletion_attack_target_vs_retain_c3ref.json"
    if neo_teacher_retain_c3_path.exists():
        delete_matching_rows(
            teacher_bound_ws,
            lambda row: len(row) >= 2 and row[0] == "neo-1.3B" and row[1] == "C6_teacher_minus_C3_teacher",
        )
    if neo_teacher_retain_c1_path.exists():
        delete_matching_rows(
            teacher_bound_ws,
            lambda row: len(row) >= 2 and row[0] == "neo-1.3B" and row[1] == "retain_pool_C6_teacher_minus_C1_teacher",
        )
    for row in neo_teacher_deletion_bound_rows():
        upsert(teacher_bound_ws, ["scale", "bound_id"], row)
    sort_sheet(teacher_bound_ws, ["scale", "bound_id"])

    teacher_kl_ws = ensure_sheet(
        wb,
        "teacher_kl",
        ["scale", "reference", "pool", "kl_c6_to_ref_mean", "kl_ref_to_c6_mean", "js_proxy_mean", "num_tokens", "file"],
    )
    for row in teacher_kl_rows():
        upsert(teacher_kl_ws, ["scale", "reference", "pool"], row)
    sort_sheet(teacher_kl_ws, ["scale", "reference", "pool"])

    teacher_feature_ws = ensure_sheet(
        wb,
        "teacher_feature_attacks",
        ["scale", "reference", "candidate_pool", "positive_class", "negative_class", "feature", "auroc", "ci_low", "ci_high", "positive_mean", "negative_mean", "n_positive", "n_negative", "file"],
    )
    for row in teacher_feature_attack_rows():
        upsert(teacher_feature_ws, ["scale", "reference", "candidate_pool", "feature"], row)
    sort_sheet(teacher_feature_ws, ["scale", "reference", "candidate_pool", "feature"])

    attack_threshold_ws = ensure_sheet(
        wb,
        "attack_thresholds",
        ["scale", "actor", "reference", "auroc", "ci_low", "ci_high", "average_precision", "tpr_at_1pct_fpr", "tpr_at_5pct_fpr", "tpr_at_10pct_fpr", "best_balanced_accuracy", "precision_at_5", "recall_at_5", "file"],
    )
    for row in participation_threshold_rows():
        upsert(attack_threshold_ws, ["scale", "actor", "reference"], row)
    sort_sheet(attack_threshold_ws, ["scale", "actor", "reference"])

    attack_gap_ws = ensure_sheet(
        wb,
        "attack_gap_stats",
        ["kind", "scale", "reference", "student_minus_teacher_auroc", "ci_low", "ci_high", "p_value", "sign_p_value", "n_positive", "n_negative", "num_seeds", "student_file", "teacher_file"],
    )
    for row in participation_gap_rows():
        upsert(attack_gap_ws, ["kind", "scale", "reference"], row)
    sort_sheet(attack_gap_ws, ["kind", "scale", "reference"])

    ref_mismatch_ws = ensure_sheet(
        wb,
        "ref_mismatch",
        ["scale", "c6_seed", "reference_kind", "reference_seed", "reference_condition", "auroc", "ci_low", "ci_high", "tpr_at_5pct_fpr", "best_balanced_accuracy", "file"],
    )
    for row in reference_mismatch_rows():
        upsert(ref_mismatch_ws, ["scale", "c6_seed", "reference_kind", "reference_seed"], row)
    sort_sheet(ref_mismatch_ws, ["scale", "reference_kind", "c6_seed", "reference_seed"])

    placebo_ws = ensure_sheet(
        wb,
        "placebo_control",
        ["status", "reason", "auroc", "ci_low", "ci_high", "file"],
    )
    if placebo_attack_rows():
        delete_matching_rows(
            placebo_ws,
            lambda row: len(row) >= 6 and row[5] == str(FOLLOWUP_DIR / "placebo_control.json"),
        )
    for row in placebo_rows():
        upsert(placebo_ws, ["status", "file"], row)
    sort_sheet(placebo_ws, ["status", "file"])

    seed_mia_1p4_ws = ensure_sheet(wb, "seed_mia_1.4B", ["seed", "condition", "auroc", "auroc_doc", "num_companies", "file"])
    seed_util_distill_1p4_ws = ensure_sheet(wb, "seed_utility_distill_1.4B", ["seed", "model", "mean_loss", "perplexity", "num_examples", "dataset", "file"])
    seed_util_holdout_1p4_ws = ensure_sheet(wb, "seed_utility_holdout_1.4B", ["seed", "model", "mean_loss", "perplexity", "num_examples", "dataset", "file"])
    pythia_seed_c6_mia_rows, pythia_seed_c6_util_distill_rows, pythia_seed_c6_util_holdout_rows = pythia_seed_c6_rows()
    for row in pythia_seed_c6_mia_rows:
        upsert(seed_mia_1p4_ws, ["seed", "condition"], row)
    for row in pythia_seed_c5_teacher_backfill_rows():
        upsert(seed_util_distill_1p4_ws, ["seed", "model"], row)
    for row in pythia_seed_c6_util_distill_rows:
        upsert(seed_util_distill_1p4_ws, ["seed", "model"], row)
    for row in pythia_seed_c6_util_holdout_rows:
        upsert(seed_util_holdout_1p4_ws, ["seed", "model"], row)
    sort_sheet(seed_mia_1p4_ws, ["seed", "condition"])
    sort_sheet(seed_util_distill_1p4_ws, ["seed", "model"])
    sort_sheet(seed_util_holdout_1p4_ws, ["seed", "model"])

    seed_mia_ws = ensure_sheet(wb, "seed_mia_neo_1.3B", ["seed", "condition", "auroc", "auroc_doc", "num_companies", "file"])
    seed_util_distill_ws = ensure_sheet(wb, "seed_utility_distill_neo_1.3B", ["seed", "model", "mean_loss", "perplexity", "num_examples", "dataset", "file"])
    seed_util_holdout_ws = ensure_sheet(wb, "seed_utility_holdout_neo_1.3B", ["seed", "model", "mean_loss", "perplexity", "num_examples", "dataset", "file"])
    seed_mia_rows, seed_util_distill_rows, seed_util_holdout_rows = seed_rows()
    for row in seed_mia_rows:
        upsert(seed_mia_ws, ["seed", "condition"], row)
    for row in neo_seed_c6_mia_rows:
        upsert(seed_mia_ws, ["seed", "condition"], row)
    for row in seed_util_distill_rows:
        upsert(seed_util_distill_ws, ["seed", "model"], row)
    for row in neo_seed_c6_utility_distill_rows:
        upsert(seed_util_distill_ws, ["seed", "model"], row)
    for row in seed_util_holdout_rows:
        upsert(seed_util_holdout_ws, ["seed", "model"], row)
    for row in neo_seed_c6_utility_holdout_rows:
        upsert(seed_util_holdout_ws, ["seed", "model"], row)
    sort_sheet(seed_mia_ws, ["seed", "condition"])
    sort_sheet(seed_util_distill_ws, ["seed", "model"])
    sort_sheet(seed_util_holdout_ws, ["seed", "model"])

    seed_run_metrics_ws = ensure_sheet(
        wb,
        "seed_run_metrics_neo_1.3B",
        [
            "seed",
            "method",
            "total_steps",
            "initial_forget_loss",
            "final_forget_loss",
            "initial_retain_loss",
            "final_retain_loss",
            "final_forget_log_ratio",
            "ref_checksum_match",
            "file",
        ],
    )
    for row in neo_seed_c6_run_metrics_rows:
        upsert(seed_run_metrics_ws, ["seed"], row)
    sort_sheet(seed_run_metrics_ws, ["seed"])

    notes_ws = ensure_sheet(wb, "notes", ["notes"])
    note = "GPT-Neo 1.3B rows synced from local_repo outputs."
    note2 = "Pythia 1.4B seeded c5 teacher retain-utility rows backfilled from Myriad outputs."
    note3 = "Pythia 1.4B C6 teacher/student rows synced from local outputs."
    note4 = "Deletion attack rows: Ref A uses per-company delta (C6 minus reference) against nonmember candidate pools; Ref B uses absolute C6 loss against nonmember candidate pools."
    note5 = "Deletion attack rows: retain-pool attacks use retained training companies as negatives under a stronger threat model; current results include C6_minus_C3 and C6_minus_C1."
    note6 = "Deletion attack rows prefixed with C6_teacher use teacher-level per-company losses; these are directly comparable to student-level retain-pool attacks."
    note7 = "Teacher attack path check: C6_teacher target losses exactly match the per-company losses in mia/c6_teacher.json; dirty-teacher reference uses teachers/c1, which is the model evaluated as C4 teacher."
    note8 = "Teacher attack structure: C6_teacher_minus_C1_teacher shifts both targets and retains upward almost equally (teacher AUROC ~0.50), while C6_teacher_minus_C3_teacher has heterogeneous target deltas (32 positive, 18 negative) against uniformly positive retain deltas (teacher AUROC ~0.08)."
    note9 = "GPT-Neo 2.7B canonical, matched, and utility rows were manually ingested from pasted Myriad outputs on 2026-04-17."
    note10 = "GPT-Neo C2/C3 rows were repaired using fixed datasets from gpt-neo-fixed-20260419; the previous duplicated GPT-Neo 2.7B C2/C3 artifacts are invalidated and should not be used."
    note11 = "Teacher KL summary: for both C1-teacher and C3-teacher references, C6 teacher diverges more on target-company tokens than on retained-company tokens, supporting a target-weighted distillation channel."
    note12 = "GPT-Neo 1.3B C6 teacher/student rows and NPO run-metrics were manually ingested from pasted Myriad outputs on 2026-04-18."
    note13 = "GPT-Neo 1.3B seeded C6 rows are being manually ingested from pasted Myriad outputs as seeds finish; seed-level NPO run metrics are tracked in seed_run_metrics_neo_1.3B."
    note14 = "Pythia Myriad C2/C3 student retry failed during distillation with non-finite active Pythia-410M logits; canonical local Pythia C2/C3 results are retained because those runs already provide valid null C2_vs_C3 comparisons."
    note15 = "GPT-Neo 1.3B C6 target-vs-retain deletion attack was ingested from Myriad outputs on 2026-04-22; retain-pool rows are labeled C6_del_retain_refC1 and C6_del_retain_refC3."
    note16 = "Pythia 1.4B seed-level C6 deletion attacks for seeds 13/17/19 are nonmember-pool stability checks only; they do not replicate the headline retain-pool Table 4 attack, whose negative class is retained training companies."
    note17 = "GPT-Neo 1.3B teacher C6 deletion attacks currently exist only for the nonmember-pool variant; C6_teacher_minus_C1_teacher AUROC is effectively the same as the Neo student nonmember-pool AUROC, so it should not be used as evidence for a teacher-student gap."
    note18 = "GPT-Neo 1.3B retain-pool/C3-reference teacher deletion attacks remain missing and are explicitly bounded in teacher_deletion_bounds because Neo teacher retain losses and C3-teacher loss files are not present."
    note19 = "Pythia 1.4B retain-pool seed replication for C6 is now complete for seeds 13/17/19; these headline retain-pool results are stored in seed_deletion_attacks_1.4B with candidate_pool=retained_companies and summarized in seed_deletion_summary_1.4B."
    note20 = f"Pythia 1.4B teacher-side retain-pool seed replication for C6 is now complete for seeds 13/17/19; teacher rows are stored separately in {SEED_TEACHER_ATTACK_SHEET} and {SEED_TEACHER_SUMMARY_SHEET}. Teacher mean-loss attacks remain weak: C6_teacher_minus_C3_teacher is near-zero/inverted across seeds, while C6_teacher_minus_C1_teacher is moderate and unstable."
    note21 = "Teacher feature attacks store canonical Pythia 1.4B retain-pool teacher-side AUROCs for richer scores than mean loss, including token-KL, high-KL token subsets, and gold-token logit-difference features, in teacher_feature_attacks."
    note22 = "GPT-Neo 1.3B retain-pool teacher deletion attacks are now computed and ingested from Myriad. The C1-reference teacher attack is strong (AUROC 0.9468), while the C3-reference teacher attack is inverted/near-zero (AUROC 0.0652), mirroring the cross-family asymmetry seen in Pythia teacher mean-loss attacks."
    note23 = "GPT-Neo 1.3B teacher KL diagnostics are now ingested from Myriad in teacher_kl; they compare token-level KL and JS-proxy divergences between C6 teacher and the C1/C3 teacher references on target and retain pools."
    note24 = "GPT-Neo 1.3B teacher feature attacks are now ingested from Myriad in teacher_feature_attacks; these are the per-company retain-pool AUROCs analogous to the canonical Pythia teacher token-KL feature attacks."
    note25 = "Participation follow-up analyses are now ingested from exp/outputs/attack_followups: attack_thresholds stores TPR@fixed-FPR and top-k metrics, attack_gap_stats stores student-teacher AUROC gap statistics, ref_mismatch stores cross-seed/same-seed reference mismatch robustness, and placebo_control records the status of the random-forget/placebo control branch."
    note26 = "Random-forget/placebo control runs are now ingested: Pythia 1.4B placebo was completed locally and GPT-Neo 1.3B placebo was completed on Myriad with local post-fetch finalization after a final aggregation bug. Placebo retain-pool attack rows are labeled C6_placebo_del_retain_refC1 and C6_placebo_del_retain_refC3 in deletion_attacks."
    note27 = "Seeded placebo student-side C1-reference retain-pool runs are ingested in seed_placebo_attacks and seed_placebo_summary when available. GPT-Neo 1.3B seeded placebo currently contributes seeds 17 and 19 from exp/outputs/myriad_seed_placebo_c1_20260428."
    stale_note9 = "GPT-Neo 2.7B canonical and matched MIA rows were manually ingested from pasted Myriad stats_bootstrap.json outputs on 2026-04-17; utility rows are still pending the actual JSON contents."
    stale_note10 = "GPT-Neo 2.7B C2 and C3 outputs are currently duplicated at the artifact level (identical MIA and utility JSONs); do not treat C2_vs_C3 at 2.7B as an independent scientific result until rerun or audited."
    stale_note15 = "GPT-Neo 1.3B C6 target-vs-retain deletion attack is still pending; current Neo C6 deletion rows use the nonmember candidate pool unless explicitly labeled target_vs_retain."
    stale_note17 = "GPT-Neo 1.3B teacher C6 deletion attacks currently exist only for the nonmember-pool variant; C6_teacher_minus_C1_teacher AUROC is effectively the same as the Neo student nonmember-pool AUROC, so it should not be used as evidence for a teacher-student gap."
    stale_note18 = "GPT-Neo 1.3B retain-pool/C3-reference teacher deletion attacks remain missing and are explicitly bounded in teacher_deletion_bounds because Neo teacher retain losses and C3-teacher loss files are not present."
    stale_note20 = "Pythia 1.4B teacher-side retain-pool seed replication for C6 is now complete for seeds 13/17/19; teacher rows are stored separately in seed_teacher_deletion_attacks_1.4B and seed_teacher_deletion_summary_1.4B. Teacher mean-loss attacks remain weak: C6_teacher_minus_C3_teacher is near-zero/inverted across seeds, while C6_teacher_minus_C1_teacher is moderate and unstable."
    neo_teacher_retain_ready = (NEO_REPAIR_1P3B_DIR / "mia_c6_teacher_deletion_attack_target_vs_retain_c3ref.json").exists() and (NEO_REPAIR_1P3B_DIR / "mia_c6_teacher_deletion_attack_target_vs_retain_c1ref.json").exists()
    notes = {notes_ws.cell(r, 1).value for r in range(2, notes_ws.max_row + 1)}
    for r in range(notes_ws.max_row, 1, -1):
        if notes_ws.cell(r, 1).value in {stale_note9, stale_note10, stale_note15, stale_note17, stale_note20}:
            notes_ws.delete_rows(r, 1)
    if neo_teacher_retain_ready:
        for r in range(notes_ws.max_row, 1, -1):
            if notes_ws.cell(r, 1).value in {note18, stale_note18}:
                notes_ws.delete_rows(r, 1)
    notes = {notes_ws.cell(r, 1).value for r in range(2, notes_ws.max_row + 1)}
    if note not in notes:
        notes_ws.append([note])
    if note2 not in notes:
        notes_ws.append([note2])
    if note3 not in notes:
        notes_ws.append([note3])
    if note4 not in notes:
        notes_ws.append([note4])
    if note5 not in notes:
        notes_ws.append([note5])
    if note6 not in notes:
        notes_ws.append([note6])
    if note7 not in notes:
        notes_ws.append([note7])
    if note8 not in notes:
        notes_ws.append([note8])
    if note9 not in notes:
        notes_ws.append([note9])
    if note10 not in notes:
        notes_ws.append([note10])
    if note11 not in notes:
        notes_ws.append([note11])
    if note12 not in notes:
        notes_ws.append([note12])
    if note13 not in notes:
        notes_ws.append([note13])
    if note14 not in notes:
        notes_ws.append([note14])
    if note15 not in notes:
        notes_ws.append([note15])
    if note16 not in notes:
        notes_ws.append([note16])
    if not neo_teacher_retain_ready and note17 not in notes:
        notes_ws.append([note17])
    if not neo_teacher_retain_ready and note18 not in notes:
        notes_ws.append([note18])
    if note19 not in notes:
        notes_ws.append([note19])
    if note20 not in notes:
        notes_ws.append([note20])
    if (PYTHIA_ATTACK_DIR / "mia_teacher_attack" / "teacher_feature_attack_summary.json").exists() and note21 not in notes:
        notes_ws.append([note21])
    if neo_teacher_retain_ready and note22 not in notes:
        notes_ws.append([note22])
    if (NEO_REPAIR_1P3B_DIR / "mia_teacher_attack" / "teacher_kl_targets_retains.json").exists() and note23 not in notes:
        notes_ws.append([note23])
    if (NEO_REPAIR_1P3B_DIR / "mia_teacher_attack" / "teacher_feature_attack_summary.json").exists() and note24 not in notes:
        notes_ws.append([note24])
    if (FOLLOWUP_DIR / "threshold_metrics.json").exists() and note25 not in notes:
        notes_ws.append([note25])
    if ((PYTHIA_PLACEBO_DIR / "mia_c6_placebo_deletion_attack_target_vs_retain.json").exists() or (NEO_PLACEBO_1P3B_DIR / "mia_c6_placebo_deletion_attack_target_vs_retain.json").exists()) and note26 not in notes:
        notes_ws.append([note26])
    if any((d / "seed_c6_placebo_deletion_attack_target_vs_retain_c1ref_summary.json").exists() for d in [PYTHIA_SEEDED_PLACEBO_DIR, NEO_SEEDED_PLACEBO_DIR]) and note27 not in notes:
        notes_ws.append([note27])

    wb.save(WORKBOOK)


if __name__ == "__main__":
    main()
